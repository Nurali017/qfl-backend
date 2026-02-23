"""
Dry-run package generator for db_translations.xlsx.

What it does:
- Reads Excel snapshot (db_translations.xlsx)
- Compares values with PostgreSQL tables by primary key `id`
- Applies comparison policies:
  - skip empty Excel cells
  - do not transliterate `*_en`
  - exclude table/id pairs from auto-update candidates (teams.id=595)
- Generates reproducible dry-run artifacts (no DB writes)

Usage:
  python dry_run_db_translations.py
  python dry_run_db_translations.py --xlsx-path /path/to/db_translations.xlsx
  python dry_run_db_translations.py --output-root /path/to/reports
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import text

from app.database import AsyncSessionLocal


CYRILLIC_RE = re.compile(r"[А-Яа-яЁёӘәҒғҚқҢңӨөҰұҮүІі]")
LATIN_RE = re.compile(r"[A-Za-z]")


@dataclass(frozen=True)
class TableConfig:
    sheet: str
    table: str
    columns: tuple[str, ...]


TABLES: tuple[TableConfig, ...] = (
    TableConfig(
        sheet="Championships",
        table="championships",
        columns=("name", "name_kz", "name_en", "short_name", "short_name_kz", "short_name_en"),
    ),
    TableConfig(
        sheet="Seasons",
        table="seasons",
        columns=("name", "name_kz", "name_en", "sponsor_name", "sponsor_name_kz"),
    ),
    TableConfig(
        sheet="Stages",
        table="stages",
        columns=("name", "name_kz", "name_en"),
    ),
    TableConfig(
        sheet="Teams",
        table="teams",
        columns=("name", "name_kz", "name_en", "city", "city_kz", "city_en"),
    ),
    TableConfig(
        sheet="Clubs",
        table="clubs",
        columns=("name", "name_kz", "name_en"),
    ),
    TableConfig(
        sheet="Players",
        table="players",
        columns=(
            "first_name",
            "first_name_kz",
            "first_name_en",
            "last_name",
            "last_name_kz",
            "last_name_en",
            "top_role",
            "top_role_kz",
            "top_role_en",
        ),
    ),
    TableConfig(
        sheet="Coaches",
        table="coaches",
        columns=(
            "first_name",
            "first_name_kz",
            "first_name_ru",
            "first_name_en",
            "last_name",
            "last_name_kz",
            "last_name_ru",
            "last_name_en",
        ),
    ),
    TableConfig(
        sheet="Referees",
        table="referees",
        columns=(
            "first_name",
            "first_name_kz",
            "first_name_ru",
            "first_name_en",
            "last_name",
            "last_name_kz",
            "last_name_ru",
            "last_name_en",
        ),
    ),
    TableConfig(
        sheet="Stadiums",
        table="stadiums",
        columns=(
            "name",
            "name_kz",
            "name_ru",
            "name_en",
            "city",
            "city_kz",
            "city_ru",
            "city_en",
            "address",
            "address_kz",
            "address_en",
        ),
    ),
    TableConfig(
        sheet="Countries",
        table="countries",
        columns=("name", "name_kz", "name_en"),
    ),
    TableConfig(
        sheet="Cities",
        table="cities",
        columns=("name", "name_kz", "name_en"),
    ),
)


# table/id pairs excluded from auto-generated SQL candidates
EXCLUDED_AUTO_UPDATE_IDS: set[tuple[str, int]] = {("teams", 595)}


def default_xlsx_path() -> Path:
    # repo root is parent of backend/
    return Path(__file__).resolve().parent.parent / "qfl-website" / "db_translations.xlsx"


def default_output_root() -> Path:
    return Path(__file__).resolve().parent / "reports"


def to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def escape_sql_literal(value: str) -> str:
    return value.replace("'", "''")


def quality_issues_for_cell(table: str, sheet: str, row_id: int, column: str, value: str) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []

    if value.startswith(" ") or value.endswith(" "):
        issues.append(
            {
                "table": table,
                "sheet": sheet,
                "id": row_id,
                "column": column,
                "issue_type": "leading_or_trailing_space",
                "value": value,
            }
        )

    if "  " in value:
        issues.append(
            {
                "table": table,
                "sheet": sheet,
                "id": row_id,
                "column": column,
                "issue_type": "double_space",
                "value": value,
            }
        )

    if column.endswith("_en"):
        has_cyr = CYRILLIC_RE.search(value) is not None
        has_lat = LATIN_RE.search(value) is not None
        if has_cyr:
            issues.append(
                {
                    "table": table,
                    "sheet": sheet,
                    "id": row_id,
                    "column": column,
                    "issue_type": "non_latin_in_en",
                    "value": value,
                }
            )
        if has_cyr and has_lat:
            issues.append(
                {
                    "table": table,
                    "sheet": sheet,
                    "id": row_id,
                    "column": column,
                    "issue_type": "mixed_latin_cyrillic_in_en",
                    "value": value,
                }
            )

    return issues


def read_excel_snapshot(xlsx_path: Path) -> dict[str, dict[int, dict[str, str]]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result: dict[str, dict[int, dict[str, str]]] = {}

    for cfg in TABLES:
        if cfg.sheet not in wb.sheetnames:
            raise ValueError(f"Missing sheet '{cfg.sheet}' in {xlsx_path}")

        ws = wb[cfg.sheet]
        headers: list[str] = []
        for col_idx in range(1, ws.max_column + 1):
            headers.append(to_text(ws.cell(1, col_idx).value).strip())

        header_to_idx = {h: i + 1 for i, h in enumerate(headers) if h}

        if "id" not in header_to_idx:
            raise ValueError(f"Sheet '{cfg.sheet}' does not contain 'id' header")

        missing_columns = [col for col in cfg.columns if col not in header_to_idx]
        if missing_columns:
            raise ValueError(f"Sheet '{cfg.sheet}' missing columns: {missing_columns}")

        sheet_rows: dict[int, dict[str, str]] = {}
        id_col = header_to_idx["id"]
        for row_idx in range(2, ws.max_row + 1):
            raw_id = ws.cell(row_idx, id_col).value
            if raw_id is None:
                continue

            try:
                row_id = int(raw_id)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"Invalid id in sheet '{cfg.sheet}' row {row_idx}: {raw_id}") from exc

            row_data: dict[str, str] = {}
            for col in cfg.columns:
                cell_value = ws.cell(row_idx, header_to_idx[col]).value
                row_data[col] = to_text(cell_value)
            sheet_rows[row_id] = row_data

        result[cfg.table] = sheet_rows

    return result


async def fetch_db_snapshot() -> dict[str, dict[int, dict[str, Any]]]:
    snapshot: dict[str, dict[int, dict[str, Any]]] = {}
    async with AsyncSessionLocal() as session:
        for cfg in TABLES:
            selected_columns = ", ".join(["id", *cfg.columns])
            rows = (
                await session.execute(text(f"SELECT {selected_columns} FROM {cfg.table}"))
            ).fetchall()

            table_map: dict[int, dict[str, Any]] = {}
            for row in rows:
                row_id = int(row[0])
                values = {cfg.columns[i]: row[i + 1] for i in range(len(cfg.columns))}
                table_map[row_id] = values
            snapshot[cfg.table] = table_map
    return snapshot


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_sql(candidates: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    lines.append("-- Auto-generated dry-run SQL candidates")
    lines.append("-- Safe pattern: UPDATE ... WHERE id = ... AND column IS DISTINCT FROM value")
    lines.append("-- NOTE: This file is generated but NOT executed by the script.")
    lines.append("")

    if not candidates:
        lines.append("-- No candidate updates for current snapshot.")
        lines.append("")
        return "\n".join(lines)

    for row in candidates:
        table = row["table"]
        column = row["column"]
        row_id = int(row["id"])
        new_value = escape_sql_literal(str(row["xlsx_value"]))
        lines.append(
            f"UPDATE {table} "
            f"SET {column} = '{new_value}' "
            f"WHERE id = {row_id} AND {column} IS DISTINCT FROM '{new_value}';"
        )
    lines.append("")
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate dry-run artifacts for db_translations.xlsx")
    parser.add_argument(
        "--xlsx-path",
        type=Path,
        default=default_xlsx_path(),
        help="Path to db_translations.xlsx",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=default_output_root(),
        help="Directory where timestamped report folder will be created",
    )
    parser.add_argument(
        "--timestamp",
        type=str,
        default=None,
        help="Optional timestamp for output folder name (format: YYYYMMDD_HHMMSS)",
    )
    return parser.parse_args()


async def run() -> None:
    args = parse_args()
    xlsx_path = args.xlsx_path.resolve()
    output_root = args.output_root.resolve()
    timestamp = args.timestamp or datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    report_dir = output_root / f"db_translations_dry_run_{timestamp}"
    report_dir.mkdir(parents=True, exist_ok=False)

    xlsx_snapshot = read_excel_snapshot(xlsx_path)
    db_snapshot = await fetch_db_snapshot()

    fill_rate_rows: list[dict[str, Any]] = []
    quality_issues: list[dict[str, Any]] = []
    candidate_updates: list[dict[str, Any]] = []
    manual_review: list[dict[str, Any]] = []
    summary_tables: dict[str, Any] = {}

    for cfg in TABLES:
        table = cfg.table
        xrows = xlsx_snapshot[table]
        drows = db_snapshot[table]
        x_ids = set(xrows.keys())
        d_ids = set(drows.keys())

        missing_in_db = sorted(x_ids - d_ids)
        extra_in_db = sorted(d_ids - x_ids)

        table_candidate_count = 0
        table_manual_count = 0

        # Fill rate + quality from Excel snapshot
        total_rows = len(xrows)
        for col in cfg.columns:
            x_non_empty = 0
            db_non_null = 0
            matched_rows = 0
            for row_id, row_data in xrows.items():
                value = row_data[col]
                if value != "":
                    x_non_empty += 1
                    quality_issues.extend(quality_issues_for_cell(table, cfg.sheet, row_id, col, value))

                db_row = drows.get(row_id)
                if db_row is not None:
                    matched_rows += 1
                    if db_row.get(col) is not None and to_text(db_row.get(col)) != "":
                        db_non_null += 1

            x_fill_pct = round((x_non_empty / total_rows * 100.0), 1) if total_rows else 0.0
            db_fill_pct = round((db_non_null / matched_rows * 100.0), 1) if matched_rows else 0.0
            fill_rate_rows.append(
                {
                    "table": table,
                    "sheet": cfg.sheet,
                    "column": col,
                    "xlsx_non_empty": x_non_empty,
                    "xlsx_total": total_rows,
                    "xlsx_fill_percent": x_fill_pct,
                    "db_non_null_on_matched_ids": db_non_null,
                    "db_matched_total": matched_rows,
                    "db_fill_percent_on_matched_ids": db_fill_pct,
                }
            )

        # Diff generation for candidates/manual review
        for row_id in sorted(x_ids):
            if row_id in missing_in_db:
                continue

            xrow = xrows[row_id]
            drow = drows[row_id]
            for col in cfg.columns:
                x_value = xrow[col]
                if x_value == "":
                    continue  # policy: do not touch empty excel cells

                db_value_raw = drow.get(col)
                db_value = to_text(db_value_raw)
                if db_value == x_value:
                    continue

                record = {
                    "table": table,
                    "sheet": cfg.sheet,
                    "id": row_id,
                    "column": col,
                    "db_value": db_value,
                    "xlsx_value": x_value,
                }

                if (table, row_id) in EXCLUDED_AUTO_UPDATE_IDS:
                    manual_review.append(
                        {
                            **record,
                            "reason": "excluded_policy_id",
                        }
                    )
                    table_manual_count += 1
                else:
                    candidate_updates.append(
                        {
                            **record,
                            "reason": "value_mismatch_non_empty_excel",
                        }
                    )
                    table_candidate_count += 1

        summary_tables[table] = {
            "sheet": cfg.sheet,
            "xlsx_rows": len(x_ids),
            "db_rows_total": len(d_ids),
            "missing_in_db_by_id_count": len(missing_in_db),
            "missing_in_db_by_id_sample": missing_in_db[:20],
            "extra_in_db_by_id_count": len(extra_in_db),
            "extra_in_db_by_id_sample": extra_in_db[:20],
            "candidate_updates_count": table_candidate_count,
            "manual_review_count": table_manual_count,
        }

    # deterministic ordering
    quality_issues.sort(key=lambda r: (r["table"], int(r["id"]), r["column"], r["issue_type"]))
    candidate_updates.sort(key=lambda r: (r["table"], int(r["id"]), r["column"]))
    manual_review.sort(key=lambda r: (r["table"], int(r["id"]), r["column"]))

    # Files
    summary_path = report_dir / "summary.json"
    fill_rate_path = report_dir / "fill_rate.csv"
    quality_path = report_dir / "quality_issues.csv"
    candidates_path = report_dir / "candidate_updates.csv"
    sql_path = report_dir / "candidate_updates.sql"
    manual_path = report_dir / "manual_review.csv"

    quality_by_type: dict[str, int] = {}
    for issue in quality_issues:
        key = issue["issue_type"]
        quality_by_type[key] = quality_by_type.get(key, 0) + 1

    summary = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "mode": "dry_run",
        "sources": {
            "xlsx_path": str(xlsx_path),
        },
        "policies": {
            "match_key": "id",
            "skip_empty_excel_cells": True,
            "transliterate_en": False,
            "excluded_auto_update_ids": [{"table": t, "id": i} for t, i in sorted(EXCLUDED_AUTO_UPDATE_IDS)],
        },
        "tables": summary_tables,
        "totals": {
            "candidate_updates_count": len(candidate_updates),
            "manual_review_count": len(manual_review),
            "quality_issues_count": len(quality_issues),
            "quality_issues_by_type": quality_by_type,
        },
        "artifacts": {
            "summary_json": str(summary_path),
            "fill_rate_csv": str(fill_rate_path),
            "quality_issues_csv": str(quality_path),
            "candidate_updates_csv": str(candidates_path),
            "candidate_updates_sql": str(sql_path),
            "manual_review_csv": str(manual_path),
        },
    }

    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    write_csv(
        fill_rate_path,
        [
            "table",
            "sheet",
            "column",
            "xlsx_non_empty",
            "xlsx_total",
            "xlsx_fill_percent",
            "db_non_null_on_matched_ids",
            "db_matched_total",
            "db_fill_percent_on_matched_ids",
        ],
        fill_rate_rows,
    )

    write_csv(
        quality_path,
        ["table", "sheet", "id", "column", "issue_type", "value"],
        quality_issues,
    )

    write_csv(
        candidates_path,
        ["table", "sheet", "id", "column", "db_value", "xlsx_value", "reason"],
        candidate_updates,
    )

    sql_path.write_text(build_sql(candidate_updates), encoding="utf-8")

    write_csv(
        manual_path,
        ["table", "sheet", "id", "column", "db_value", "xlsx_value", "reason"],
        manual_review,
    )

    print(f"Report directory: {report_dir}")
    print("Coverage summary:")
    for cfg in TABLES:
        table = cfg.table
        info = summary_tables[table]
        print(
            f"  {table}: missing_in_db_by_id={info['missing_in_db_by_id_count']}, "
            f"extra_in_db_by_id={info['extra_in_db_by_id_count']}"
        )
    print("Issue summary:")
    print(f"  quality_issues={len(quality_issues)}")
    print(f"  candidate_updates={len(candidate_updates)}")
    print(f"  manual_review={len(manual_review)}")


if __name__ == "__main__":
    asyncio.run(run())
